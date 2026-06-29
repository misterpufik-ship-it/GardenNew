# -*- coding: utf-8 -*-
"""Generate data.json from bundled Lounge xlsx files."""
import json
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "config.json"
OUT_PATH = ROOT / "data.json"

REPORT_FILE = ROOT / "отчет Никита.xlsx"
STATEMENT_FILE = ROOT / "Лаундж" / "выписка апрель 2026.xlsx"


def norm_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def iso_date(d):
    return d.isoformat() if d else None


def to_num(v):
    if v is None or v == "":
        return None
    try:
        return round(float(str(v).replace(" ", "").replace(",", ".")), 2)
    except (TypeError, ValueError):
        return None


def rub(n):
    return int(abs(float(n or 0)))


def cell_val(ws, r, c):
    if not c or c < 1:
        return None
    cell = ws.cell(r, c)
    return cell.value


def is_bn(v, filt="бн"):
    return str(v or "").strip().lower() == filt.lower()


def read_report_lounge(wb, cfg):
    sc = cfg["lounge"]["report"]
    name = wb.sheetnames[sc["sheetIndex"]]
    ws = wb[name]
    out = []
    current_date = None
    uid = 0
    for r in range(1, ws.max_row + 1):
        maybe = norm_date(cell_val(ws, r, sc["dateCol"]))
        if maybe:
            current_date = maybe
        sm = to_num(cell_val(ws, r, sc["sumCol"]))
        if not is_bn(cell_val(ws, r, sc["paymentCol"])) or sm is None:
            continue
        uid += 1
        out.append({
            "id": f"e{uid}",
            "date": iso_date(current_date),
            "sheet": name,
            "category": str(cell_val(ws, r, sc["categoryCol"]) or "").strip(),
            "amount": sm,
            "comment": str(cell_val(ws, r, sc["commentCol"]) or "").strip(),
            "status": "unmatched",
            "groupNum": None,
            "matchId": None,
        })
    return out


def read_statement_lounge(wb, cfg):
    sc = cfg["lounge"]["statement"]
    name = wb.sheetnames[sc["sheetIndex"]]
    ws = wb[name]
    out = []
    current_date = None
    uid = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, 4):
            raw = cell_val(ws, r, c)
            if raw is None:
                continue
            text = str(raw)
            if "Дата:" in text:
                import re
                m = re.search(r"(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})", text)
                if m:
                    parts = m.group(1).replace("-", ".").replace("/", ".").split(".")
                    y = int(parts[2])
                    if y < 100:
                        y += 2000
                    current_date = date(y, int(parts[1]), int(parts[0]))
                break
            d = norm_date(raw)
            if d and len(text) < 20:
                current_date = d
                break
        debit = to_num(cell_val(ws, r, sc["debitCol"]))
        if debit is None or debit <= 0:
            continue
        uid += 1
        out.append({
            "id": f"s{uid}",
            "date": iso_date(current_date),
            "opNum": str(cell_val(ws, r, sc["opNumCol"]) or "").strip(),
            "counterparty": str(cell_val(ws, r, sc["counterpartyCol"]) or "").strip(),
            "purpose": str(cell_val(ws, r, sc["purposeCol"]) or "").strip(),
            "amount": debit,
            "status": "unmatched",
            "groupNum": None,
            "matchId": None,
        })
    return out


def find_subset(items, target, min_len):
    result = None

    def dfs(idx, sm, picked):
        nonlocal result
        if result:
            return
        if sm == target and len(picked) >= min_len:
            result = picked[:]
            return
        if sm > target or idx >= len(items):
            return
        dfs(idx + 1, sm, picked)
        picked.append(items[idx])
        dfs(idx + 1, sm + rub(items[idx]["amount"]), picked)
        picked.pop()

    dfs(0, 0, [])
    return result


def reconcile(expenses, statements):
    used_exp, used_stmt = set(), set()
    matches = []
    group_num = 1
    expenses = deepcopy(expenses)
    statements = deepcopy(statements)

    for exp in sorted(expenses, key=lambda x: x.get("date") or ""):
        if exp["id"] in used_exp:
            continue
        cands = [s for s in statements if s["id"] not in used_stmt and rub(s["amount"]) == rub(exp["amount"])]
        cands.sort(key=lambda s: 0 if s.get("date") == exp.get("date") else 1)
        if not cands:
            continue
        stmt = cands[0]
        mid = f"m{len(matches)+1}"
        used_exp.add(exp["id"])
        used_stmt.add(stmt["id"])
        exp.update(status="exact", matchId=mid)
        stmt.update(status="exact", matchId=mid)
        matches.append({"id": mid, "type": "exact", "groupNum": None, "expenseIds": [exp["id"]], "statementId": stmt["id"]})

    by_date = {}
    for stmt in statements:
        if stmt["id"] in used_stmt:
            continue
        by_date.setdefault(stmt.get("date") or "_", []).append(stmt)

    for dt, stmts in by_date.items():
        for stmt in stmts:
            if stmt["id"] in used_stmt:
                continue
            day_exp = [e for e in expenses if e["id"] not in used_exp and (e.get("date") or "_") == dt]
            if len(day_exp) < 2:
                continue
            combo = find_subset(day_exp, rub(stmt["amount"]), 2)
            if not combo:
                continue
            mid = f"m{len(matches)+1}"
            g = group_num
            group_num += 1
            for exp in combo:
                used_exp.add(exp["id"])
                exp.update(status="group", groupNum=g, matchId=mid)
            used_stmt.add(stmt["id"])
            stmt.update(status="group", groupNum=g, matchId=mid)
            matches.append({
                "id": mid,
                "type": "group",
                "groupNum": g,
                "expenseIds": [e["id"] for e in combo],
                "statementId": stmt["id"],
            })

    stats = {
        "exact": sum(1 for m in matches if m["type"] == "exact"),
        "group": sum(1 for m in matches if m["type"] == "group"),
        "unmatchedExp": sum(1 for e in expenses if e["status"] == "unmatched"),
        "unmatchedStmt": sum(1 for s in statements if s["status"] == "unmatched"),
    }
    return expenses, statements, matches, stats


def build_pair_rows(expenses, statements, matches):
    exp_map = {e["id"]: e for e in expenses}
    stmt_map = {s["id"]: s for s in statements}
    rows = []
    link_num = 0

    for m in matches:
        link_num += 1
        stmt = stmt_map[m["statementId"]]
        for i, eid in enumerate(m["expenseIds"]):
            exp = exp_map[eid]
            rows.append({
                "linkNum": link_num if i == 0 else None,
                "matchId": m["id"],
                "status": m["type"],
                "groupNum": m.get("groupNum"),
                "matchType": "Точное" if m["type"] == "exact" else "Группа",
                "groupSize": len(m["expenseIds"]),
                "isFirstInGroup": i == 0,
                "expDate": exp.get("date"),
                "category": exp.get("category"),
                "reportAmount": exp.get("amount"),
                "comment": exp.get("comment"),
                "sheet": exp.get("sheet"),
                "stmtDate": stmt.get("date") if i == 0 else None,
                "opNum": stmt.get("opNum") if i == 0 else None,
                "counterparty": stmt.get("counterparty") if i == 0 else None,
                "statementAmount": stmt.get("amount") if i == 0 else None,
                "purpose": stmt.get("purpose") if i == 0 else None,
            })

    for exp in expenses:
        if exp["status"] != "unmatched":
            continue
        link_num += 1
        rows.append({
            "linkNum": link_num,
            "matchId": None,
            "status": "unmatched",
            "groupNum": None,
            "matchType": "Не найдено",
            "groupSize": 1,
            "isFirstInGroup": True,
            "expDate": exp.get("date"),
            "category": exp.get("category"),
            "reportAmount": exp.get("amount"),
            "comment": exp.get("comment"),
            "sheet": exp.get("sheet"),
            "stmtDate": None,
            "opNum": None,
            "counterparty": None,
            "statementAmount": None,
            "purpose": None,
        })

    for stmt in statements:
        if stmt["status"] != "unmatched":
            continue
        link_num += 1
        rows.append({
            "linkNum": link_num,
            "matchId": None,
            "status": "unmatched_stmt",
            "groupNum": None,
            "matchType": "Не найдено",
            "groupSize": 1,
            "isFirstInGroup": True,
            "expDate": None,
            "category": None,
            "reportAmount": None,
            "comment": None,
            "sheet": None,
            "stmtDate": stmt.get("date"),
            "opNum": stmt.get("opNum"),
            "counterparty": stmt.get("counterparty"),
            "statementAmount": stmt.get("amount"),
            "purpose": stmt.get("purpose"),
        })

    rows.sort(key=lambda r: (r.get("expDate") or r.get("stmtDate") or "", r.get("linkNum") or 9999))
    return rows


def main():
    cfg = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    report_wb = openpyxl.load_workbook(REPORT_FILE, data_only=True)
    stmt_wb = openpyxl.load_workbook(STATEMENT_FILE, data_only=True)
    expenses = read_report_lounge(report_wb, cfg)
    statements = read_statement_lounge(stmt_wb, cfg)
    expenses, statements, matches, stats = reconcile(expenses, statements)
    detail = build_pair_rows(expenses, statements, matches)

    data = {
        "entity": "lounge",
        "reportFile": REPORT_FILE.name,
        "statementFile": STATEMENT_FILE.name,
        "expenses": expenses,
        "statements": statements,
        "matches": matches,
        "stats": stats,
        "pairRows": detail,
    }
    OUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Saved {OUT_PATH}")
    print("Stats:", stats)
    print("Expenses:", len(expenses), "Statements:", len(statements), "Pair rows:", len(detail))


if __name__ == "__main__":
    main()
