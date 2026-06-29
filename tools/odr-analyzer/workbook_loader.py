"""Load workbook and classify sheets."""
from __future__ import annotations

import re
from pathlib import Path
from openpyxl import load_workbook

PATTERNS = {
    'odr': re.compile(r'^одр$', re.I),
    'daily_moyka': re.compile(r'^адмир$', re.I),
    'daily_akku': re.compile(r'^уделка$', re.I),
    'writeoffs': re.compile(r'^списания', re.I),
    'stock': re.compile(r'^остатки\s+на\s+склад', re.I),
    'transfers': re.compile(r'^перемещения', re.I),
    'inv_bar_moyka': re.compile(r'^адмирал\s', re.I),
    'inv_bar_akku': re.compile(r'^аккуратова\s', re.I),
    'inv_km_moyka': re.compile(r'^км\s+мойка', re.I),
    'inv_km_akku': re.compile(r'^км\s+аккуратова', re.I),
    'inv_kitchen_moyka': re.compile(r'^кухня\s+мойка', re.I),
    'inv_kitchen_akku': re.compile(r'^кухня\s+аккуратова', re.I),
}


def load_odr_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    sheets = {}
    sheet_names = []
    for ws in wb.worksheets:
        name = ws.title.strip()
        sheet_names.append({'name': name, 'rows': ws.max_row, 'cols': ws.max_column})
        for key, pat in PATTERNS.items():
            if pat.search(name):
                sheets[key] = ws
    return {'workbook': wb, 'sheets': sheets, 'sheet_names': sheet_names}
